from openai import OpenAI
import os
import json
import openpyxl
import xlsxwriter

textdata = openpyxl.load_workbook('DepressionText_parsed.xlsx')

# Get all the sheets in the workbook
sheets = textdata.sheetnames

#agelist: age, text
samples = []

# # Iterate through the sheets
# for sheet in sheets:
#     # Get the active worksheet
#     ws = textdata[sheet]
#     # Iterate through the rows in the worksheet
#     for row in ws.rows:
#         print(row[1].value , " " ,row[0].value)
#         samples.append([row[1].value,row[0].value])

# Iterate through specific sheet
ws = textdata[sheets[4]]
for row in ws.rows:
    #print(row[1].value , " " ,row[0].value)
    samples.append([row[1].value,row[0].value])

results = [['age','length']];
#d = predicted age - actual age. (aiming for 0)

for i in samples:

    if(i[0] == "age"):
        continue
    
    results.append([i[0],len(i[1])]);


workbook = xlsxwriter.Workbook('DataArchive.xlsx');
worksheet = workbook.add_worksheet("01282024-0")

r = 0;
c = 0;

for value in (results):
    print(value)
    worksheet.write(r,c,value[0]);
    worksheet.write(r,c+1,value[1]);
    r += 1

workbook.close()