import openpyxl
import pandas as pd

textdata = openpyxl.load_workbook('Depression_Text.xlsx')
td = textdata.active

#age range from 13 - 17
with pd.ExcelWriter("DepressionText_parsedLength.xlsx") as writer:
    for iter in range (1,5):
        # get all the data within the age
        text = []
        label = []
        age = []
        gender = []
        for i in range (2 , td.max_row+1):
            if td.cell(row=i, column = 3).value != None:  #if age data exist,
                length = len(td.cell(row=i, column = 1).value)
                if length <= iter*50 and length > (iter-1)*50 or iter == 4 and length >151:
                    text.append(td.cell(row=i, column = 1).value)
                    label.append(td.cell(row=i, column = 2).value)
                    age.append(td.cell(row=i, column = 3).value)
                    gender.append(td.cell(row=i, column = 4).value)

        #create datasheet
        df = pd.DataFrame({'text':text,'label':label,'age':age,'gender':gender})
        if(iter == 4):
            df.to_excel(writer,sheet_name="150+ Data",index=False)
        else:
            df.to_excel(writer,sheet_name=str((iter-1)*50)+"-"+str(iter*50)+"Data",index=False)


    

                



        



