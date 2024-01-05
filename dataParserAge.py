import openpyxl
import pandas as pd

textdata = openpyxl.load_workbook('Depression_Text.xlsx')
td = textdata.active

#age range from 13 - 17
#parsing the depression text database based on age.
with pd.ExcelWriter("DepressionText_parsed.xlsx") as writer:
    for ageiter in range (13,18):
        # get all the data within the age
        text = []
        gender = []
        for i in range (2 , td.max_row+1):
            if td.cell(row=i, column = 3).value != None:  #if age data exist,
                age = int(td.cell(row=i, column = 3).value)
                if age == ageiter:
                    text.append(td.cell(row=i, column = 1).value)
                    gender.append(td.cell(row=i, column = 4).value)

        #create datasheet
        df = pd.DataFrame({'text':text,'gender':gender})
        df.to_excel(writer,sheet_name=str(ageiter)+"Data",index=False)