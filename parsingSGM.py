import openpyxl
import pandas as pd

textdata = openpyxl.load_workbook('studentWritingSample.xlsx')
td = textdata.active

#age range from 13 - 17
with pd.ExcelWriter("studentWritingSample_parsedLength.xlsx") as writer:
    for iter in range (1,5):
        # get all the data within the age
        text = []
        age = []
        for i in range (2 , td.max_row+1):
            length = len(td.cell(row=i, column = 1).value)
            if length <= iter*50 and length > (iter-1)*50 or iter == 4 and length >151:
                text.append(td.cell(row=i, column = 1).value)
                age.append(td.cell(row=i, column = 2).value)

        #create datasheet
        df = pd.DataFrame({'text':text,'age':age})
        if(iter == 4):
            df.to_excel(writer,sheet_name="150+ Data",index=False)
        else:
            df.to_excel(writer,sheet_name=str((iter-1)*50)+"-"+str(iter*50)+"Data",index=False)


    

                



        



